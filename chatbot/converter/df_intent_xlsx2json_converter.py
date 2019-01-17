import openpyxl

class Conversation:

    def __init__(self, contentType, contentName, question, answer):
        self.contentType = contentType
        self.contentName = contentName
        self.question = question
        self.answer = answer

    def  __str__(self):
        return "질문: " + self.question + "\n답변: " + self.answer + "\n"


### 엑셀파일 경로 입력 ###
file_path = 'C:/df/test/test_1210.xlsx'
wb = openpyxl.load_workbook(file_path)

#ws = wb.get_sheet_by_name('intent')
"""
print('file_path : ', file_path)
print('wb : ', wb)
print('wb.get_sheet_names() : ', wb.get_sheet_names())
print('sheet : ', ws)
print('sheet[1] : ', ws[1])
print('sheet : ', ws.cell(row = 1, column = 3).value)

print('ws.iter_rows() ; ', ws.iter_rows())
print('len(ws.iter_rows()) : ', len(ws.iter_rows()))


for row in ws.iter_rows():
    for cell in row:
        print(cell)
"""

# .active : 활성화된 시트를 불러옴
ws = wb.active
conversations = []

for r in ws.rows:
    c = Conversation(r[0].value, r[1].value, r[2].value, r[3].value)
    conversations.append(c)

for c in conversations:
    print(str(c))

print('총 ', len(conversations), '개의 대화가 존재합니다.')

wb.close()

output_file_path = 'C:/df/intent_result/'

prev = str(conversations[0].contentType) + "-" + str(conversations[0].contentName)

# .json 1행 답변파일의 생성
f = open(output_file_path + prev + '.json', 'w', encoding= 'UTF-8')
f.write('{ "id": "", "name": "' + prev + '", "auto": true, "contexts": [], "responses": [ { "resetContexts": false, "affectedContexts": [], "parameters": [], "messages": [ { "type": 0, "lang": "ko", "speech": "' + conversations[0]. answer + '" } ], "defaultResponsePlatforms": {}, "speech": [] } ], "priority": 500000, "webhookUsed": false, "webhookForSlotFilling": false, "fallbackIntent": false, "events": [] }')
f.close()
# .json 1행 답변파일의 작성완료

# .json 유저say파일 생성
f = open(output_file_path + prev + '_usersays_ko.json', 'w', encoding= 'UTF-8')
f.write("[")
f.write('{ "id": "", "data": [ { "text": "' + conversations[0].question + '", "userDefined": false } ], "isTemplate": false, "count": 0 }')


i = 1

while True:
    if i >= len(conversations):
        f.write("]")
        f.close()
        break;

    c = conversations[i]

    if prev == str(c.contentType) + "-" + str(c.contentName):
        f.write(',{ "id": "", "data": [ { "text": "' + c.question + '", "userDefined": false } ], "isTemplate": false, "count": 0 }')
    else:
        f.write("]")
        f.close()

        prev = str(c.contentType) + "-" + str(c.contentName)
        f = open(output_file_path + prev + '.json', 'w', encoding='UTF-8')
        f.write('{"id": "", "name": "' + prev + '", "auto": true, "contexts": [], "responses": [ { "resetContexts": false, "affectedContexts": [], "parameters": [], "messages": [ { "type": 0, "lang": "ko", "speech": "' + c.answer + '" } ], "defaultResponsePlatforms": {}, "speech": [] } ], "priority": 500000, "webhookUsed": false, "webhookForSlotFilling": false, "fallbackIntent": false, "events": [] }')
        f.close()

        f = open(output_file_path + prev + '_usersays_ko.json', 'w', encoding='UTF-8')
        f.write("[")
        f.write('{ "id": "", "data": [ { "text": "' + c.question + '", "userDefined": false } ], "isTemplate": false, "count": 0 },')

    i = i + 1


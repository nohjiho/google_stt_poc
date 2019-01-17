import openpyxl
#코드를 보며 자부심을 느끼기 위해 카피라이터를 프로그램 상단에 붙인다. ㅋㅋ
#entityconverter 와 같은 클래스명은
#첫글자 대문자 이며, EntityConverter와 같이 의미별 앞글자를 대문자로.

class entityconverter :
    #entityName 와 같은  지역변수 entity_name와 같이 헝가리안표기법
    #function명은 헝가리언 표기법
    # 띄어쓰기는 한칸만
    def __init__(self, entityName, entityrep, synonyms) :
        self.entityName = entityName
        self.entityrep = entityrep
        self.synonyms = synonyms

    #펑션마다 펑션에 대한 설명을 주석으로 단다.
    def __str__(self) :
        return "entity: " + self.entityName + "\n유의어: " + self.synonyms + "\n"

### 엑셀파일 경로 입력 ###
file_path = 'C:/df/convert test_entity.xlsx'
entxl = openpyxl.load_workbook(file_path)

#ws = wb.get_sheet_by_name('intent')
"""
print('file_path : ', file_path)
print('wb : ', wb)
print('wb.get_sheet_names() : ', wb.get_sheet_names())
print('sheet : ', ws)
print('sheet[1] : ', ws[1])
print('sheet : ', ws.cell(row = 1, column = 3).value)

print('ws.iter_rows() ; ', ws.iter_rows())
print('len(ws.iter_rows()) : ', len(ws.iter_rows()))


for row in ws.iter_rows():
    for cell in row:
        print(cell)
"""

# .active : 활성화된 시트를 불러옴
entityst = entxl.active #변수에 대한 주석은 잘 달지 않지만, 주요한 변수에는 단다.
synonyms = []

for r in entityst.rows :
    e = entityconverter(r[0].value, r[1].value, r[2].value)
    synonyms.append(e)

for e in synonyms :
    print(str(e))

print('총 ', len(synonyms), '개의 대화가 존재합니다.')
#중요
entxl.close()

output_file_path = 'C:/df/entity_result/'
prev = str(synonyms[0].entityName)
print(prev)

# .json 1행 답변파일의 생성
f = open(output_file_path + prev + '.json', 'w', encoding= 'UTF-8')
f.write('{  "id": "","name": "'+ prev +'","isOverridable": true,"isEnum": false,"automatedExpansion": false}')
f.close()
# .json 1행 답변파일의 작성완료
# "id": ""를 공란으로 둘 것. 공란으로 만들지 않고 restore, import를 진행하면 복수의 entity중 한 개만 등록이 됨.

strsyn = str(synonyms[0].synonyms)
synlist = strsyn.split(',')
print(synlist)


# .json _엔트리스_ko파일 생성
f = open(output_file_path + prev + '_entries_ko.json', 'w', encoding= 'UTF-8')
f.write("[")
f.write('{ "value": "' + synonyms[0].entityrep + '", "synonyms": ' + str(synlist) + '}')

i = 1

# 뭘하는 루프인지. 주석을 달아준다.
while True:
    if i >= len(synonyms):
        f.write("]")
        f.close()
        break;

    e = synonyms[i]
    strsyn = str(e.synonyms)
    synlist = strsyn.split(',')

    if prev == str(e.entityName):
        f.write(',{"value": "' + e.entityrep + '", "synonyms": ' + str(synlist) + '}')
    else:
        f.write("]")
        f.close()

        prev = str(e.entityName)
        f = open(output_file_path + prev + '.json', 'w', encoding='UTF-8')
        f.write('{  "id": "","name": "'
                + prev +
                '","isOverridable": true,"isEnum": false,"automatedExpansion": false}')
        f.close()

        f = open(output_file_path + prev + '_entries_ko.json', 'w', encoding='UTF-8')
        f.write("[")
        f.write('{"value": "'
                + e.entityrep +
                '", "synonyms": '
                + str(synlist) +
                '}')

    i = i + 1